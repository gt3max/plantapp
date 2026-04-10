"""
Toxicity v4 — deep enrichment from TPPT xlsx + Duke's CSV + NC State scrape.

Step 1: TPPT full xlsx (845 plants, 1587 toxins, toxic_parts, severity)
Step 2: Duke's Phytochemical CSV (plant→chemical with LD50/toxic activities)
Step 3: NC State toxicity tags scrape (5,000+ plants)

Usage:
    python3 toxicity_v4_deep.py --step 1          # TPPT only
    python3 toxicity_v4_deep.py --step 2          # Duke's only
    python3 toxicity_v4_deep.py --step 3 --limit 100  # NC State limited
    python3 toxicity_v4_deep.py --dry-run
"""
import sys
import os
import re
import csv
import json
import time
import urllib.request

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from pathlib import Path
env_path = Path(__file__).parent / '.env'
if env_path.exists():
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            key, val = line.split('=', 1)
            os.environ.setdefault(key.strip(), val.strip())

from turso_sync import turso_query, turso_batch, store_source_data

UA = 'PlantApp/1.0 (plantapp.pro)'
TPPT_XLSX = os.path.join(os.path.dirname(__file__), 'data', 'tppt_full.xlsx')
DUKE_CSV_DIR = os.path.join(os.path.dirname(__file__), 'data', 'duke_csv')


# =========================================================
# STEP 1: TPPT full xlsx
# =========================================================
def step1_tppt(dry_run=False):
    """Extract full toxicity from TPPT xlsx."""
    print(f"\n=== STEP 1: TPPT full xlsx ===", flush=True)

    try:
        import openpyxl
    except ImportError:
        print("  ERROR: pip install openpyxl", flush=True)
        return

    wb = openpyxl.load_workbook(TPPT_XLSX, read_only=True)

    # Plants sheet
    ws_plants = wb['Plants']
    plants = []
    header = [c.value for c in next(ws_plants.iter_rows(min_row=1, max_row=1))]
    for row in ws_plants.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header, row))
        if d.get('Latin_plant_name'):
            plants.append(d)
    print(f"  TPPT plants: {len(plants)}", flush=True)

    # Phytotoxins sheet
    ws_toxins = wb['Phytotoxins']
    toxins = {}
    header_t = [c.value for c in next(ws_toxins.iter_rows(min_row=1, max_row=1))]
    for row in ws_toxins.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header_t, row))
        if d.get('Phytotoxin_number'):
            toxins[d['Phytotoxin_number']] = d

    # Relationships sheet
    ws_rels = wb['Relationships']
    rels = {}  # plant_number → [toxin_names]
    header_r = [c.value for c in next(ws_rels.iter_rows(min_row=1, max_row=1))]
    for row in ws_rels.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header_r, row))
        pn = d.get('Plant_number', '')
        tn = d.get('Phytotoxin_name', '')
        if pn and tn:
            if pn not in rels:
                rels[pn] = []
            rels[pn].append(tn)

    # Toxicity_data sheet (LD50)
    ws_ld50 = wb['Toxicity_data']
    ld50 = {}  # toxin_number → [{animal, route, dose}]
    header_l = [c.value for c in next(ws_ld50.iter_rows(min_row=1, max_row=1))]
    for row in ws_ld50.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header_l, row))
        tn = d.get('Phytotoxin_number', '')
        if tn:
            if tn not in ld50:
                ld50[tn] = []
            ld50[tn].append(d)

    wb.close()
    print(f"  Toxins: {len(toxins)}, Relationships: {sum(len(v) for v in rels.values())}, LD50: {sum(len(v) for v in ld50.values())}", flush=True)

    # Match with our DB
    our = turso_query("SELECT plant_id, scientific FROM plants WHERE scientific IS NOT NULL")
    our_by_name = {}
    for p in our:
        sci = p['scientific'].lower()
        our_by_name[sci] = p['plant_id']
        parts = sci.split()
        if len(parts) >= 2:
            our_by_name[' '.join(parts[:2])] = p['plant_id']

    stmts = []
    stats = {'matched': 0, 'toxic_set': 0, 'parts_set': 0, 'toxins_set': 0}

    for plant in plants:
        latin = (plant.get('Latin_plant_name') or '').strip()
        # Clean: remove author names after species
        latin_clean = re.sub(r'\s*\(.*?\)', '', latin).strip()
        latin_lower = latin_clean.lower()
        parts = latin_lower.split()

        pid = our_by_name.get(latin_lower)
        if not pid and len(parts) >= 2:
            pid = our_by_name.get(' '.join(parts[:2]))
        if not pid:
            continue

        stats['matched'] += 1
        pn = plant.get('Plant_number', '')
        human_tox = plant.get('Human_toxicity') or ''
        animal_tox = plant.get('Animal_toxicity') or ''
        toxic_part = plant.get('Toxic_plant_part') or ''
        plant_toxin_names = rels.get(pn, [])

        if not dry_run:
            # Store raw TPPT data
            fields = {}
            if human_tox:
                fields['human_toxicity'] = str(human_tox)
            if animal_tox:
                fields['animal_toxicity'] = str(animal_tox)
            if toxic_part:
                fields['toxic_plant_part'] = str(toxic_part)
            if plant_toxin_names:
                fields['toxin_names'] = ', '.join(sorted(set(plant_toxin_names)))[:300]

            store_source_data(pid, 'tppt_v2', fields)

            # Update care
            is_toxic = human_tox and any(w in str(human_tox).lower() for w in ['toxic', 'strong', 'very', 'lethal', 'highly', '++', '+++'])
            if is_toxic:
                stmts.append(("UPDATE care SET toxic_to_humans = 1 WHERE plant_id = ?", [pid]))
                stats['toxic_set'] += 1

            is_animal_toxic = animal_tox and any(w in str(animal_tox).lower() for w in ['toxic', 'strong', 'lethal', '++', '+++'])
            if is_animal_toxic:
                stmts.append(("UPDATE care SET toxic_to_pets = 1 WHERE plant_id = ? AND (toxic_to_pets = 0 OR toxic_to_pets IS NULL)", [pid]))

            if toxic_part:
                stmts.append(("UPDATE care SET toxic_parts = ? WHERE plant_id = ? AND (toxic_parts IS NULL OR toxic_parts = '')",
                              [str(toxic_part)[:200], pid]))
                stats['parts_set'] += 1

            if plant_toxin_names:
                toxin_str = ', '.join(sorted(set(plant_toxin_names)))[:200]
                stmts.append(("UPDATE care SET toxicity_note = ? WHERE plant_id = ? AND (toxicity_note IS NULL OR toxicity_note = '')",
                              [f'Contains: {toxin_str}', pid]))
                stats['toxins_set'] += 1

            # Severity from human_toxicity text
            severity = None
            ht = str(human_tox).lower()
            if '+++' in ht or 'very' in ht or 'lethal' in ht or 'highly' in ht:
                severity = 'Severe'
            elif '++' in ht or 'strong' in ht or 'toxic' in ht:
                severity = 'Moderate'
            elif '+' in ht or 'weak' in ht:
                severity = 'Mild'
            if severity:
                stmts.append(("UPDATE care SET toxicity_severity = ? WHERE plant_id = ? AND (toxicity_severity IS NULL OR toxicity_severity = '')",
                              [severity, pid]))

            if len(stmts) >= 100:
                turso_batch(stmts); stmts = []

    if stmts and not dry_run:
        turso_batch(stmts)

    print(f"  Matched: {stats['matched']}, toxic: {stats['toxic_set']}, parts: {stats['parts_set']}, toxins: {stats['toxins_set']}", flush=True)


# =========================================================
# STEP 2: Duke's Phytochemical CSV
# =========================================================
def step2_duke(dry_run=False):
    """Parse Duke's CSV — find plants with known toxic chemicals."""
    print(f"\n=== STEP 2: Duke's Phytochemical CSV ===", flush=True)

    # Known toxic activity keywords
    TOXIC_ACTIVITIES = {
        'Toxic', 'Poison', 'Pesticide', 'Cytotoxic', 'Hepatotoxic',
        'Nephrotoxic', 'Neurotoxic', 'Cardiotoxic', 'Abortifacient',
        'Teratogenic', 'Mutagenic', 'Carcinogenic', 'Emetic', 'Purgative',
        'LD50', 'Lethal',
    }

    # Parse CHEMICALS.csv → chemical_id → chemical_name
    chemicals = {}
    chem_path = os.path.join(DUKE_CSV_DIR, 'CHEMICALS.csv')
    if os.path.exists(chem_path):
        with open(chem_path, encoding='latin-1') as f:
            for row in csv.DictReader(f):
                cid = row.get('CHEM_ID', '').strip()
                name = row.get('CHEMICAL', '').strip()
                if cid and name:
                    chemicals[cid] = name

    print(f"  Duke chemicals: {len(chemicals)}", flush=True)

    # Parse ACTIVITIES.csv → find toxic activities per chemical
    toxic_chemicals = set()
    act_path = os.path.join(DUKE_CSV_DIR, 'ACTIVITIES.csv')
    if os.path.exists(act_path):
        with open(act_path, encoding='latin-1') as f:
            for row in csv.DictReader(f):
                activity = row.get('ACTIVITY', '').strip()
                cid = row.get('CHEM_ID', '').strip()
                if cid and any(t.lower() in activity.lower() for t in TOXIC_ACTIVITIES):
                    toxic_chemicals.add(cid)

    print(f"  Chemicals with toxic activities: {len(toxic_chemicals)}", flush=True)

    # Parse FARMACY_NEW.csv → plant → chemical relationships
    plant_toxics = {}  # scientific_name → set(toxic_chemical_names)
    farm_path = os.path.join(DUKE_CSV_DIR, 'FARMACY_NEW.csv')
    if not os.path.exists(farm_path):
        farm_path = os.path.join(DUKE_CSV_DIR, 'FARMACY.csv')

    if os.path.exists(farm_path):
        with open(farm_path, encoding='latin-1') as f:
            for row in csv.DictReader(f):
                plant = row.get('FNFTAX', '').strip().lower()
                cid = row.get('CHEM_ID', '').strip()
                if plant and cid in toxic_chemicals:
                    if plant not in plant_toxics:
                        plant_toxics[plant] = set()
                    chem_name = chemicals.get(cid, cid)
                    plant_toxics[plant].add(chem_name)

    print(f"  Plants with toxic chemicals: {len(plant_toxics)}", flush=True)

    # Match with our DB
    our = turso_query("SELECT plant_id, scientific FROM plants WHERE scientific IS NOT NULL")
    our_by_name = {}
    for p in our:
        sci = p['scientific'].lower()
        our_by_name[sci] = p['plant_id']
        parts = sci.split()
        if len(parts) >= 2:
            our_by_name[' '.join(parts[:2])] = p['plant_id']

    stmts = []
    matched = 0
    for plant_name, tox_chems in plant_toxics.items():
        pid = our_by_name.get(plant_name)
        if not pid:
            parts = plant_name.split()
            if len(parts) >= 2:
                pid = our_by_name.get(' '.join(parts[:2]))

        if pid:
            matched += 1
            if not dry_run:
                chems_str = ', '.join(sorted(tox_chems))[:300]
                stmts.append((
                    "INSERT OR REPLACE INTO source_data (plant_id, source, field, value, fetched_at) VALUES (?, 'duke', 'toxic_chemicals', ?, datetime('now'))",
                    [pid, chems_str]
                ))
                stmts.append(("UPDATE care SET toxic_to_humans = 1 WHERE plant_id = ? AND (toxic_to_humans = 0 OR toxic_to_humans IS NULL)",
                              [pid]))
                stmts.append(("UPDATE care SET toxicity_note = ? WHERE plant_id = ? AND (toxicity_note IS NULL OR toxicity_note = '')",
                              [f'Contains toxic compounds: {chems_str[:150]}', pid]))

                if len(stmts) >= 100:
                    turso_batch(stmts); stmts = []

    if stmts and not dry_run:
        turso_batch(stmts)

    print(f"  Matched with our DB: {matched}", flush=True)


# =========================================================
# STEP 3: NC State toxicity scrape
# =========================================================
def step3_ncstate(dry_run=False, limit=None):
    """Scrape NC State toxicity tags for our plants."""
    print(f"\n=== STEP 3: NC State toxicity scrape ===", flush=True)

    # Get plants we have in NC State source_data but no toxicity
    plants = turso_query("""
        SELECT DISTINCT sd.plant_id, p.scientific FROM source_data sd
        JOIN plants p ON sd.plant_id = p.plant_id
        WHERE sd.source = 'ncstate'
        AND sd.plant_id NOT IN (
            SELECT plant_id FROM source_data WHERE source = 'ncstate_toxicity'
        )
    """)

    if limit:
        plants = plants[:limit]

    print(f"  NC State plants to check: {len(plants)}", flush=True)

    stmts = []
    stats = {'checked': 0, 'toxic_found': 0, 'principle': 0, 'error': 0}

    for i, plant in enumerate(plants):
        pid = plant['plant_id']
        sci = plant.get('scientific', '')

        # Build URL from plant_id
        url_slug = pid.replace('_', '-')
        url = f'https://plants.ces.ncsu.edu/plants/{url_slug}/'

        try:
            req = urllib.request.Request(url, headers={'User-Agent': UA})
            with urllib.request.urlopen(req, timeout=15) as resp:
                html = resp.read().decode()

            stats['checked'] += 1

            # Extract toxicity fields
            principle = re.findall(r'Toxic[^<]*Principle[^<]*:.*?<span[^>]*>([^<]+)', html, re.DOTALL)
            part = re.findall(r'Poison[^<]*Part[^<]*:.*?<span[^>]*>([^<]+)', html, re.DOTALL)
            severity = re.findall(r'Poison[^<]*Severity[^<]*:.*?<span[^>]*>([^<]+)', html, re.DOTALL)
            symptoms = re.findall(r'Poison[^<]*Symptom[^<]*:.*?<span[^>]*>([^<]+)', html, re.DOTALL)

            has_toxicity = principle or part or severity or symptoms

            if has_toxicity and not dry_run:
                stats['toxic_found'] += 1

                if principle:
                    stmts.append((
                        "INSERT OR REPLACE INTO source_data (plant_id, source, field, value, fetched_at) VALUES (?, 'ncstate_toxicity', 'toxic_principle', ?, datetime('now'))",
                        [pid, principle[0][:200]]
                    ))
                    stmts.append(("UPDATE care SET toxicity_note = ? WHERE plant_id = ? AND (toxicity_note IS NULL OR toxicity_note = '')",
                                  [principle[0][:200], pid]))
                    stmts.append(("UPDATE care SET toxic_to_humans = 1 WHERE plant_id = ?", [pid]))
                    stats['principle'] += 1

                if part:
                    stmts.append((
                        "INSERT OR REPLACE INTO source_data (plant_id, source, field, value, fetched_at) VALUES (?, 'ncstate_toxicity', 'poison_part', ?, datetime('now'))",
                        [pid, part[0][:200]]
                    ))
                    stmts.append(("UPDATE care SET toxic_parts = ? WHERE plant_id = ? AND (toxic_parts IS NULL OR toxic_parts = '')",
                                  [part[0][:200], pid]))

                if severity:
                    sev = severity[0].strip()
                    sev_map = {'Low': 'Mild', 'Medium': 'Moderate', 'High': 'Severe'}
                    mapped = sev_map.get(sev, sev)
                    stmts.append((
                        "INSERT OR REPLACE INTO source_data (plant_id, source, field, value, fetched_at) VALUES (?, 'ncstate_toxicity', 'poison_severity', ?, datetime('now'))",
                        [pid, sev]
                    ))
                    stmts.append(("UPDATE care SET toxicity_severity = ? WHERE plant_id = ? AND (toxicity_severity IS NULL OR toxicity_severity = '')",
                                  [mapped, pid]))

                if symptoms:
                    stmts.append((
                        "INSERT OR REPLACE INTO source_data (plant_id, source, field, value, fetched_at) VALUES (?, 'ncstate_toxicity', 'poison_symptoms', ?, datetime('now'))",
                        [pid, symptoms[0][:300]]
                    ))
                    stmts.append(("UPDATE care SET toxicity_symptoms = ? WHERE plant_id = ? AND (toxicity_symptoms IS NULL OR toxicity_symptoms = '')",
                                  [symptoms[0][:300], pid]))

                if len(stmts) >= 50:
                    turso_batch(stmts); stmts = []

            elif not has_toxicity and not dry_run:
                # NC State page exists but no toxicity → likely safe
                stmts.append((
                    "INSERT OR REPLACE INTO source_data (plant_id, source, field, value, fetched_at) VALUES (?, 'ncstate_toxicity', 'no_toxicity_listed', 'true', datetime('now'))",
                    [pid]
                ))

        except Exception as e:
            stats['error'] += 1

        if (i + 1) % 100 == 0:
            print(f"  [{i+1}/{len(plants)}] checked={stats['checked']} toxic={stats['toxic_found']} principle={stats['principle']} err={stats['error']}", flush=True)

        time.sleep(0.5)

    if stmts and not dry_run:
        turso_batch(stmts)

    print(f"  Done: checked={stats['checked']}, toxic={stats['toxic_found']}, principle={stats['principle']}, error={stats['error']}", flush=True)


def show_results():
    """Show toxicity stats."""
    print(f"\n=== TOXICITY STATUS ===", flush=True)
    tox = turso_query("SELECT toxic_to_humans, COUNT(*) as c FROM care GROUP BY toxic_to_humans")
    for t in tox:
        print(f"  toxic_to_humans={str(t['toxic_to_humans']):5s}: {t['c']:>6}", flush=True)

    for f in ['toxic_parts', 'toxicity_severity', 'toxicity_symptoms', 'toxicity_note']:
        r = turso_query(f"SELECT COUNT(*) as c FROM care WHERE {f} IS NOT NULL AND {f} != ''")
        print(f"  {f}: {r[0]['c']}", flush=True)

    sources = turso_query("""
        SELECT source, COUNT(DISTINCT plant_id) as c FROM source_data
        WHERE source IN ('tppt_v2','duke','ncstate_toxicity','calpoison','efsa','wikipedia_toxic','pfaf_toxicity','aspca','tppt','cbif','family_toxicity_inference')
        GROUP BY source ORDER BY c DESC
    """)
    print(f"\nSources:", flush=True)
    for s in sources:
        print(f"  {s['source']:30s} {s['c']:>5}", flush=True)


if __name__ == '__main__':
    dry_run = '--dry-run' in sys.argv
    step_only = None
    limit = None

    if '--step' in sys.argv:
        idx = sys.argv.index('--step')
        if idx + 1 < len(sys.argv):
            step_only = int(sys.argv[idx + 1])
    if '--limit' in sys.argv:
        idx = sys.argv.index('--limit')
        if idx + 1 < len(sys.argv):
            limit = int(sys.argv[idx + 1])

    if step_only is None or step_only == 1:
        step1_tppt(dry_run)
    if step_only is None or step_only == 2:
        step2_duke(dry_run)
    if step_only is None or step_only == 3:
        step3_ncstate(dry_run, limit)

    show_results()
